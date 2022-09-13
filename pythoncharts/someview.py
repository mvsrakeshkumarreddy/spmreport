# HttpResponse is used to
# pass the information
# back to view
import openpyxl
#import datetime
#from matplotlib import pyplot as plt
from django.template import Context
import time
import os
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
#import numpy as np
import io, base64, uuid
from openpyxl import workbook,load_workbook
from io import BytesIO
from datetime import datetime, timedelta, timezone, tzinfo
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.template import loader
from plotly.offline import plot
import plotly.graph_objs as go


@login_required(login_url = '/login')
def something (request) :
    print(request.method)
    if request.method == "GET":
        return render(request, "someindex.html")
    else:
        def time_diff(hr2,min2,sec2,hr1,min1,sec1):
            time_diff = (hr2-hr1) + ((min2-min1)/60) + ((sec2-sec1)/3600)
            return time_diff
        #sir here in in the variable some it is giving the formula present in the excel sheet so by just adding "data_only=True" in wb1 and wb2 we can get the value of the cell instead of formula
        
        #datas = request.POST
        #print(datas)
        #signal_sample=datas.get("signalsample")

        #signal_speed=datas.get("signalspeed")


        signal_sample = request.FILES["signalsample"]
        signal_speed = request.FILES["signalspeed"]

        #wb1 = load_workbook(filename = 'C:\wamp64\www\env_site\pythoncharts\pythoncharts\signal speed.xlsx',data_only=True)
        #wb2 = load_workbook(filename = 'C:\wamp64\www\env_site\pythoncharts\pythoncharts\signal sample.xlsx',data_only=True)
        
        wb1 = load_workbook(filename = request.FILES['signalspeed'].file,data_only=True)
        wb2 = load_workbook(filename = request.FILES['signalsample'].file,data_only=True)

        #wb1 = openpyxl.load_workbook(signal_sample)
        #wb2 = openpyxl.load_workbook(signal_speed)

        sh1 = wb1['Sheet1'] 

        sh2 = wb2['Sheet1']


        row1_ct = sh1.max_row

        col1_ct = sh1.max_column

        row2_ct = sh2.max_row

        col2_ct = sh2.max_column

        speed_list = []

        signal_list = []

        row_num_dist2 = 2

        column_num_signal_dist2 = 6

        cumu_dist1 = 2

        j = 2

        print(row1_ct)

        print(row2_ct)


        for i in range(2,row1_ct+1):
            #sir here the value that is being compared is giving some none values which cannot be compared with any of the datatype so i just added a if statement 
            
            someright = (sh1.cell(i,cumu_dist1).value)
            
            some = (sh2.cell(row_num_dist2,column_num_signal_dist2).value)
            
            if someright != None and some != None:
                if (some) == (someright):
                    
                    speed_list.append(sh1.cell(i,1).value)
                    
                    signal_list.append(sh2.cell(row_num_dist2,1).value)
                    
                    i = i+1
                    
                    row_num_dist2 = row_num_dist2 + 1
                
                elif (some) > (someright):
                    
                    i = i+1
                
                elif (some) < (someright):
                    
                    hr2 = sh1.cell(i,3).value.hour
                    min2 = sh1.cell(i,3).value.minute
                    sec2 = sh1.cell(i,3).value.second
                    
                    hr1 = sh1.cell(i-1,3).value.hour
                    min1 = sh1.cell(i-1,3).value.minute
                    sec1 = sh1.cell(i-1,3).value.second
                    
                    time_dif = time_diff(hr2,min2,sec2,hr1,min1,sec1)
                    
                    speed_dif = sh1.cell(i,1).value-sh1.cell(i-1,1).value
                    
                    dist_dif = (sh1.cell(i,2).value-sh1.cell(i-1,2).value)/1000
                    
                    distan_travelled = ((some)-(sh1.cell(i-1,cumu_dist1).value))/1000

                    #print(speed_dif)
                    #print(time_dif)
                    
                    #Sir this is to check how many decimal points are available in the variable
                    timedifcheck = str(time_dif)
                    #print(timedifcheck)
                    noofdecimal = timedifcheck[::-1].find('.')
                    #print(noofdecimal)
                    #print("{:.19f}".format(round(speed_dif, 2)))
                    
                    #Sir here we are getting an error division by zero, so i have given this if statement
                    if time_dif!=0:
                        
                        accel = speed_dif/time_dif
                    
                    x = (sh1.cell(i-1,1).value)**2
                    
                    y = (2*accel*distan_travelled)
                    
                    z = (x+y)**0.5
                    
                    speed_list.append(z)
                    
                    signal_list.append(sh2.cell(row_num_dist2,1).value)
                    
                    i = i+1
                    
                    row_num_dist2 = row_num_dist2 + 1
                
            
    #time_diff = sh1.cell(3,3).value-sh1.cell(2,3).value

    #print(time_diff)

    #time_diff_in_hours = (time_diff.hour)+((time_diff.minute)/60)+((time_diff.second)/3600)

    #print(time_diff_in_hours)
            
            
        print(speed_list)        
                
        print(signal_list)

        x = signal_list

        y = speed_list
        
        xaxis = []
        for i in range(len(signal_list)):
            xaxis.append(i)
        
        graphsize = (len(signal_list))/2
        if graphsize>100:
            graphsize = 90

        #plt.xticks(xaxis,x,rotation=90)

        """
        plt.figure(figsize=(graphsize,8))

        plt.plot(xaxis,y,color = 'green',linestyle = 'solid',linewidth = 2,marker = 'o',markerfacecolor = 'blue',markersize = 9)
        #plt.xticks(rotation=90)
        plt.xticks(xaxis, signal_list, rotation =90)
        plt.tick_params(labelsize=6)

        plt.xlabel('signal locations')

        plt.ylabel('speed graph')

        plt.title('My first graph')

        #plt.plot(x,y,color = 'green',linestyle = 'solid',linewidth = 2,marker = 'o',markerfacecolor = 'blue',markersize = 9)

        #plt.show()
        buffer = BytesIO()
        plt.savefig(buffer, format = 'png')
        buffer.seek(0)
        image_png = buffer.getvalue()
        #chart.clear()
        chart = base64.b64encode(image_png)
        chart = chart.decode('utf-8')
        buffer.close()
        plt.clf()
        """



        fig = go.Figure()
        scatter = go.Scatter(x=xaxis, y=speed_list, mode='lines', name='test', opacity=0.8, marker_color='blue')
        fig.add_trace(scatter)
        #fig.update_xaxes(type='category')
        fig.update_layout(xaxis = dict(tickmode = 'array',tickvals = xaxis, ticktext = signal_list))
        plot_div = plot(fig, output_type='div')
        #plot_div = plot([Scatter(x=signal_list, y=speed_list, mode='lines', name='test', opacity=0.8, marker_color='blue')], output_type='div')

        #return render(request, "index.html", context={'plot_div': plot_div})

        
        return render(request, "someindex.html", {"y" : y, "x" : x, 'plot_div': plot_div})
